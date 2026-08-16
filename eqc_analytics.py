"""
EQC Google Analytics 4 - Data Fetcher + Excel Export  v2.0
===========================================================
Requer: google-analytics-data, google-auth, openpyxl
Instalar: pip install google-analytics-data google-auth openpyxl
"""

import os
import sys
from datetime import date, timedelta, datetime
from google.oauth2 import service_account
from google.analytics.data_v1beta import BetaAnalyticsDataClient
from google.analytics.data_v1beta.types import (
    RunReportRequest, DateRange, Dimension, Metric, OrderBy,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

# CONFIG
SERVICE_ACCOUNT_FILE = r"C:\eqc\credentials.json"
GA4_PROPERTY_ID      = "530030780"
XLSX_PATH            = r"C:\eqc\Analytics\EQC_Analytics.xlsx"

# DATAS INCREMENTAIS
DATE_END = date.today().strftime("%Y-%m-%d")
RUN_DATE = datetime.now().strftime("%Y-%m-%d %H:%M")

def _get_last_run_date():
    if not os.path.exists(XLSX_PATH):
        return None
    try:
        wb = load_workbook(XLSX_PATH, data_only=True)
        if "History" not in wb.sheetnames:
            return None
        ws = wb["History"]
        last_date = None
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row and row[2] is not None:
                val = str(row[2]).strip()[:10]
                try:
                    d = date.fromisoformat(val)
                    if last_date is None or d > last_date:
                        last_date = d
                except ValueError:
                    pass
        return last_date
    except Exception:
        return None

_last = _get_last_run_date()
if _last is not None:
    DATE_START = (_last + timedelta(days=1)).strftime("%Y-%m-%d")
else:
    DATE_START = (date.today() - timedelta(days=30)).strftime("%Y-%m-%d")

_days      = max((date.fromisoformat(DATE_END) - date.fromisoformat(DATE_START)).days + 1, 1)
PREV_END   = (date.fromisoformat(DATE_START) - timedelta(days=1)).strftime("%Y-%m-%d")
PREV_START = (date.fromisoformat(PREV_END) - timedelta(days=_days - 1)).strftime("%Y-%m-%d")

# CORES EQC
BG_DARK   = "050F1E"
BG_MID    = "0E2035"
GOLD      = "C9A84C"
BLUE_GREY = "8899BB"
WHITE     = "FFFFFF"
GREEN     = "3DBA5E"
RED_COL   = "E05555"
HEADER_BG = "0A1A30"

# HELPERS ESTILO
def _fill(h): return PatternFill("solid", start_color=h, end_color=h)
def _font(hex_color=WHITE, bold=False, size=10, name="Courier New"):
    return Font(color=hex_color, bold=bold, size=size, name=name)
def _border():
    s = Side(style="thin", color="0E2035")
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center")
def _left():   return Alignment(horizontal="left",   vertical="center")
def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def style_header(ws, row, col, text, bg=HEADER_BG, fg=GOLD, size=10, bold=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font=_font(fg,bold=bold,size=size); c.fill=_fill(bg)
    c.alignment=_center(); c.border=_border(); return c

def style_data(ws, row, col, value, bg=BG_DARK, fg=WHITE, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font=_font(fg); c.fill=_fill(bg)
    c.alignment=_left() if align=="left" else _center()
    c.border=_border(); return c

def fmt_delta(cur, prev):
    try:
        c, p = float(cur), float(prev)
        if p == 0: return ("NEW",100) if c>0 else ("--",0)
        pct = ((c-p)/abs(p))*100
        return (f"+{pct:.1f}%" if pct>=0 else f"{pct:.1f}%"), pct
    except: return "--", 0

def sheet_title(ws, row, ncols, text):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=ncols)
    c=ws.cell(row=row,column=1,value=text)
    c.font=_font(GOLD,bold=True,size=11); c.fill=_fill(BG_DARK)
    c.alignment=_center(); ws.row_dimensions[row].height=24

def build_table_sheet(ws, title, headers, rows, col_widths, value_cols=None):
    ws.sheet_view.showGridLines=False
    ncols=len(headers); sheet_title(ws,1,ncols,title)
    for ci,h in enumerate(headers,1): style_header(ws,2,ci,h)
    ws.row_dimensions[2].height=18
    for i,w in enumerate(col_widths,1): set_col_width(ws,i,w)
    for ri,row_data in enumerate(rows,3):
        bg=BG_DARK if ri%2==1 else BG_MID
        for ci,val in enumerate(row_data,1):
            fg=GOLD if (value_cols and ci in value_cols) else WHITE
            style_data(ws,ri,ci,val,bg=bg,fg=fg,
                       align="center" if (value_cols and ci in value_cols) else "left")
        ws.row_dimensions[ri].height=16

# AUTENTICACAO
def get_client():
    creds=service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE,
        scopes=["https://www.googleapis.com/auth/analytics.readonly"])
    return BetaAnalyticsDataClient(credentials=creds)

# FETCH GA4
def _run(client,start,end,dimensions,metrics,order_bys=None,limit=20):
    req=RunReportRequest(
        property=f"properties/{GA4_PROPERTY_ID}",
        date_ranges=[DateRange(start_date=start,end_date=end)],
        dimensions=[Dimension(name=d) for d in dimensions],
        metrics=[Metric(name=m) for m in metrics],
        order_bys=order_bys or [], limit=limit)
    return client.run_report(req)

def _f(v):
    try: return float(v.value)
    except: return 0.0

def fetch_overview(client, start, end):
    req=RunReportRequest(
        property=f"properties/{GA4_PROPERTY_ID}",
        date_ranges=[DateRange(start_date=start,end_date=end)],
        metrics=[Metric(name=m) for m in [
            "sessions","totalUsers","newUsers","screenPageViews",
            "averageSessionDuration","bounceRate","engagementRate",
            "sessionsPerUser","screenPageViewsPerSession"]])
    try:
        r=client.run_report(req).rows[0].metric_values
        nu=int(_f(r[2])); tot=int(_f(r[1]))
        return {"sessions":int(_f(r[0])),"users":tot,"new_users":nu,
                "returning_users":max(tot-nu,0),"pageviews":int(_f(r[3])),
                "avg_dur":round(_f(r[4]),1),"bounce":round(_f(r[5])*100,1),
                "engagement_rate":round(_f(r[6])*100,1),
                "sessions_per_user":round(_f(r[7]),2),
                "pages_per_session":round(_f(r[8]),2)}
    except:
        return {k:0 for k in ["sessions","users","new_users","returning_users",
                               "pageviews","avg_dur","bounce","engagement_rate",
                               "sessions_per_user","pages_per_session"]}

def fetch_top_pages(client):
    resp=_run(client,DATE_START,DATE_END,["pagePath"],
        ["screenPageViews","totalUsers","averageSessionDuration","bounceRate","engagementRate"],
        [OrderBy(metric=OrderBy.MetricOrderBy(metric_name="screenPageViews"),desc=True)],10)
    return [(r.dimension_values[0].value,
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             round(_f(r.metric_values[2]),1),round(_f(r.metric_values[3])*100,1),
             round(_f(r.metric_values[4])*100,1)) for r in resp.rows]

def fetch_landing_pages(client):
    resp=_run(client,DATE_START,DATE_END,["landingPage"],
        ["sessions","totalUsers","bounceRate","engagementRate","averageSessionDuration"],
        [OrderBy(metric=OrderBy.MetricOrderBy(metric_name="sessions"),desc=True)],10)
    return [(r.dimension_values[0].value,
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             round(_f(r.metric_values[2])*100,1),round(_f(r.metric_values[3])*100,1),
             round(_f(r.metric_values[4]),1)) for r in resp.rows]

def fetch_channels(client):
    resp=_run(client,DATE_START,DATE_END,["sessionDefaultChannelGrouping"],
        ["sessions","totalUsers","newUsers","engagementRate","bounceRate","averageSessionDuration"],
        [OrderBy(metric=OrderBy.MetricOrderBy(metric_name="sessions"),desc=True)],10)
    return [(r.dimension_values[0].value,
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),int(_f(r.metric_values[2])),
             round(_f(r.metric_values[3])*100,1),round(_f(r.metric_values[4])*100,1),
             round(_f(r.metric_values[5]),1)) for r in resp.rows]

def fetch_sources(client):
    resp=_run(client,DATE_START,DATE_END,["sessionSource","sessionMedium"],
        ["sessions","totalUsers","engagementRate","averageSessionDuration"],
        [OrderBy(metric=OrderBy.MetricOrderBy(metric_name="sessions"),desc=True)],15)
    return [(r.dimension_values[0].value,r.dimension_values[1].value,
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             round(_f(r.metric_values[2])*100,1),round(_f(r.metric_values[3]),1)) for r in resp.rows]

def fetch_countries(client):
    resp=_run(client,DATE_START,DATE_END,["country"],
        ["totalUsers","sessions","engagementRate","averageSessionDuration"],
        [OrderBy(metric=OrderBy.MetricOrderBy(metric_name="totalUsers"),desc=True)],15)
    return [(r.dimension_values[0].value,
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             round(_f(r.metric_values[2])*100,1),round(_f(r.metric_values[3]),1)) for r in resp.rows]

def fetch_cities(client):
    resp=_run(client,DATE_START,DATE_END,["city","country"],
        ["totalUsers","sessions","engagementRate"],
        [OrderBy(metric=OrderBy.MetricOrderBy(metric_name="totalUsers"),desc=True)],15)
    return [(r.dimension_values[0].value,r.dimension_values[1].value,
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             round(_f(r.metric_values[2])*100,1)) for r in resp.rows]

def fetch_devices(client):
    resp=_run(client,DATE_START,DATE_END,["deviceCategory"],
        ["sessions","totalUsers","engagementRate","averageSessionDuration"],
        [OrderBy(metric=OrderBy.MetricOrderBy(metric_name="sessions"),desc=True)])
    return [(r.dimension_values[0].value,
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             round(_f(r.metric_values[2])*100,1),round(_f(r.metric_values[3]),1)) for r in resp.rows]

def fetch_browsers(client):
    resp=_run(client,DATE_START,DATE_END,["browser"],
        ["sessions","totalUsers","engagementRate"],
        [OrderBy(metric=OrderBy.MetricOrderBy(metric_name="sessions"),desc=True)],10)
    return [(r.dimension_values[0].value,
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             round(_f(r.metric_values[2])*100,1)) for r in resp.rows]

def fetch_os(client):
    resp=_run(client,DATE_START,DATE_END,["operatingSystem"],
        ["sessions","totalUsers","engagementRate"],
        [OrderBy(metric=OrderBy.MetricOrderBy(metric_name="sessions"),desc=True)],10)
    return [(r.dimension_values[0].value,
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             round(_f(r.metric_values[2])*100,1)) for r in resp.rows]

def fetch_events(client):
    resp=_run(client,DATE_START,DATE_END,["eventName"],
        ["eventCount","totalUsers","eventCountPerUser"],
        [OrderBy(metric=OrderBy.MetricOrderBy(metric_name="eventCount"),desc=True)],15)
    return [(r.dimension_values[0].value,
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             round(_f(r.metric_values[2]),2)) for r in resp.rows]

def fetch_new_vs_returning(client):
    resp=_run(client,DATE_START,DATE_END,["newVsReturning"],
        ["totalUsers","sessions","engagementRate","averageSessionDuration","screenPageViewsPerSession"],
        [OrderBy(metric=OrderBy.MetricOrderBy(metric_name="totalUsers"),desc=True)])
    L={"new":"New","returning":"Returning"}
    return [(L.get(r.dimension_values[0].value,r.dimension_values[0].value),
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             round(_f(r.metric_values[2])*100,1),round(_f(r.metric_values[3]),1),
             round(_f(r.metric_values[4]),2)) for r in resp.rows]

def fetch_weekly(client):
    start=(date.today()-timedelta(weeks=12)).strftime("%Y-%m-%d")
    resp=_run(client,start,date.today().strftime("%Y-%m-%d"),["year","week"],
        ["sessions","totalUsers","engagementRate"],
        [OrderBy(dimension=OrderBy.DimensionOrderBy(dimension_name="year")),
         OrderBy(dimension=OrderBy.DimensionOrderBy(dimension_name="week"))],52)
    return [(r.dimension_values[0].value+"-W"+r.dimension_values[1].value,
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             round(_f(r.metric_values[2])*100,1)) for r in resp.rows]

def fetch_monthly(client):
    start=(date.today()-timedelta(days=365)).strftime("%Y-%m-%d")
    resp=_run(client,start,date.today().strftime("%Y-%m-%d"),["year","month"],
        ["sessions","totalUsers","newUsers","screenPageViews","bounceRate","engagementRate"],
        [OrderBy(dimension=OrderBy.DimensionOrderBy(dimension_name="year")),
         OrderBy(dimension=OrderBy.DimensionOrderBy(dimension_name="month"))],24)
    return [(r.dimension_values[0].value+"-"+r.dimension_values[1].value.zfill(2),
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             int(_f(r.metric_values[2])),int(_f(r.metric_values[3])),
             round(_f(r.metric_values[4])*100,1),round(_f(r.metric_values[5])*100,1)) for r in resp.rows]

def fetch_day_of_week(client):
    resp=_run(client,DATE_START,DATE_END,["dayOfWeek"],
        ["sessions","totalUsers","engagementRate"],
        [OrderBy(dimension=OrderBy.DimensionOrderBy(dimension_name="dayOfWeek"))],7)
    DAYS=["Sunday","Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"]
    return [(DAYS[int(r.dimension_values[0].value)],
             int(_f(r.metric_values[0])),int(_f(r.metric_values[1])),
             round(_f(r.metric_values[2])*100,1)) for r in resp.rows]

def fetch_hour_of_day(client):
    resp=_run(client,DATE_START,DATE_END,["hour"],
        ["sessions","engagementRate"],
        [OrderBy(dimension=OrderBy.DimensionOrderBy(dimension_name="hour"))],24)
    return [(r.dimension_values[0].value.zfill(2)+":00",
             int(_f(r.metric_values[0])),
             round(_f(r.metric_values[1])*100,1)) for r in resp.rows]

# DASHBOARD
def build_dashboard(ws, ov, prev_ov):
    ws.sheet_view.showGridLines=False

    ws.merge_cells("A1:L1")
    c=ws["A1"]; c.value="EQC  |  GOOGLE ANALYTICS 4  |  evolutionquantumcapital.com"
    c.font=_font(GOLD,bold=True,size=13); c.fill=_fill(BG_DARK); c.alignment=_center()
    ws.row_dimensions[1].height=30

    ws.merge_cells("A2:L2")
    c=ws["A2"]
    c.value=(f"PERIOD: {DATE_START} -> {DATE_END}  |  "
             f"PREV: {PREV_START} -> {PREV_END}  |  GENERATED: {RUN_DATE}")
    c.font=_font(BLUE_GREY,size=9); c.fill=_fill(BG_DARK); c.alignment=_center()
    ws.row_dimensions[2].height=16; ws.row_dimensions[3].height=6

    # KPIs linha 1: 6 x 2 cols = 12
    kpis1=[
        ("SESSIONS",      ov["sessions"],         prev_ov["sessions"],         f"{ov['sessions']:,}"),
        ("UNIQUE USERS",  ov["users"],             prev_ov["users"],             f"{ov['users']:,}"),
        ("NEW USERS",     ov["new_users"],         prev_ov["new_users"],         f"{ov['new_users']:,}"),
        ("PAGE VIEWS",    ov["pageviews"],         prev_ov["pageviews"],         f"{ov['pageviews']:,}"),
        ("BOUNCE RATE",  -ov["bounce"],           -prev_ov["bounce"],            f"{ov['bounce']}%"),
        ("AVG DURATION",  ov["avg_dur"],           prev_ov["avg_dur"],           f"{int(ov['avg_dur'])}s"),
    ]
    for i,(label,cur,prev,display) in enumerate(kpis1):
        sc=i*2+1; sl=get_column_letter(sc); el=get_column_letter(sc+1)
        ws.merge_cells(f"{sl}4:{el}4")
        c=ws[f"{sl}4"]; c.value=label; c.font=_font(BLUE_GREY,size=9)
        c.fill=_fill(HEADER_BG); c.alignment=_center(); c.border=_border()
        ws.merge_cells(f"{sl}5:{el}5")
        c=ws[f"{sl}5"]; c.value=display; c.font=_font(GOLD,bold=True,size=16)
        c.fill=_fill(BG_MID); c.alignment=_center(); c.border=_border()
        ws.merge_cells(f"{sl}6:{el}6")
        c=ws[f"{sl}6"]; ds,dp=fmt_delta(cur,prev); c.value=f"vs prev  {ds}"
        c.font=_font(GREEN if dp>=0 else RED_COL,size=8,bold=True)
        c.fill=_fill(BG_DARK); c.alignment=_center(); c.border=_border()
    ws.row_dimensions[4].height=16; ws.row_dimensions[5].height=28
    ws.row_dimensions[6].height=14; ws.row_dimensions[7].height=6

    # KPIs linha 2: 4 x 3 cols = 12
    kpis2=[
        ("ENGAGEMENT RATE", ov["engagement_rate"],   prev_ov["engagement_rate"],  f"{ov['engagement_rate']}%"),
        ("RETURNING USERS", ov["returning_users"],   prev_ov["returning_users"],  f"{ov['returning_users']:,}"),
        ("SESSIONS / USER", ov["sessions_per_user"], prev_ov["sessions_per_user"],str(ov["sessions_per_user"])),
        ("PAGES / SESSION", ov["pages_per_session"], prev_ov["pages_per_session"],str(ov["pages_per_session"])),
    ]
    for i,(label,cur,prev,display) in enumerate(kpis2):
        sc=i*3+1; sl=get_column_letter(sc); el=get_column_letter(sc+2)
        ws.merge_cells(f"{sl}8:{el}8")
        c=ws[f"{sl}8"]; c.value=label; c.font=_font(BLUE_GREY,size=9)
        c.fill=_fill(HEADER_BG); c.alignment=_center(); c.border=_border()
        ws.merge_cells(f"{sl}9:{el}9")
        c=ws[f"{sl}9"]; c.value=display; c.font=_font(GOLD,bold=True,size=14)
        c.fill=_fill(BG_MID); c.alignment=_center(); c.border=_border()
        ws.merge_cells(f"{sl}10:{el}10")
        c=ws[f"{sl}10"]; ds,dp=fmt_delta(cur,prev); c.value=f"vs prev  {ds}"
        c.font=_font(GREEN if dp>=0 else RED_COL,size=8,bold=True)
        c.fill=_fill(BG_DARK); c.alignment=_center(); c.border=_border()
    ws.row_dimensions[8].height=16; ws.row_dimensions[9].height=24
    ws.row_dimensions[10].height=14; ws.row_dimensions[11].height=6

    for i in range(1,13): set_col_width(ws,i,13)

# HISTORY
def build_history_sheet(ws, ov):
    headers=["RUN DATE","PERIOD START","PERIOD END","SESSIONS","USERS","NEW USERS",
             "RETURNING","PAGE VIEWS","AVG DUR (s)","BOUNCE %","ENGAGEMENT %",
             "SESS/USER","PAGES/SESS"]
    col_widths=[18,14,14,12,10,12,12,12,12,12,14,12,12]

    if ws.max_row<=1:
        ws.sheet_view.showGridLines=False
        ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
        c=ws["A1"]; c.value="EQC  |  ANALYTICS RUN HISTORY"
        c.font=_font(GOLD,bold=True,size=11); c.fill=_fill(BG_DARK); c.alignment=_center()
        ws.row_dimensions[1].height=24
        for ci,h in enumerate(headers,1): style_header(ws,2,ci,h)
        ws.row_dimensions[2].height=18
        for i,w in enumerate(col_widths,1): set_col_width(ws,i,w)

    for row in ws.iter_rows(min_row=3,values_only=True):
        if row and str(row[2]).strip()[:10]==DATE_END:
            print(f"  [History] {DATE_END} ja existe. A ignorar duplicado.")
            return

    nr=ws.max_row+1; bg=BG_DARK if nr%2==1 else BG_MID
    data=[RUN_DATE,DATE_START,DATE_END,
          ov["sessions"],ov["users"],ov["new_users"],ov["returning_users"],
          ov["pageviews"],ov["avg_dur"],ov["bounce"],ov["engagement_rate"],
          ov["sessions_per_user"],ov["pages_per_session"]]
    for ci,val in enumerate(data,1):
        fg=GOLD if ci>3 else BLUE_GREY
        c=ws.cell(row=nr,column=ci,value=val)
        c.font=_font(fg); c.fill=_fill(bg); c.alignment=_center(); c.border=_border()
        if isinstance(val,int): c.number_format='#,##0'
        elif isinstance(val,float): c.number_format='#,##0.0'
    ws.row_dimensions[nr].height=16

# CHARTS
def add_bar(ws,title,dws,sr,er,lc,vc,anchor,w=14,h=10):
    chart=BarChart(); chart.type="col"; chart.title=title; chart.style=10
    chart.add_data(Reference(dws,min_col=vc,min_row=sr-1,max_row=er),titles_from_data=True)
    chart.set_categories(Reference(dws,min_col=lc,min_row=sr,max_row=er))
    chart.width=w; chart.height=h; ws.add_chart(chart,anchor)

def add_line(ws,title,dws,sr,er,lc,vc,anchor,w=18,h=10):
    chart=LineChart(); chart.title=title; chart.style=10
    chart.add_data(Reference(dws,min_col=vc,min_row=sr-1,max_row=er),titles_from_data=True)
    chart.set_categories(Reference(dws,min_col=lc,min_row=sr,max_row=er))
    chart.width=w; chart.height=h; ws.add_chart(chart,anchor)

# TECHNOLOGY SHEET
def build_technology_sheet(ws, devices, browsers, os_data):
    ws.sheet_view.showGridLines=False; row=1
    sections=[
        ("DEVICES",           ["DEVICE","SESSIONS","USERS","ENGAGEMENT %","AVG DURATION (s)"],devices,  [30,12,10,14,18]),
        ("BROWSERS",          ["BROWSER","SESSIONS","USERS","ENGAGEMENT %"],                  browsers, [30,12,10,14]),
        ("OPERATING SYSTEMS", ["OS","SESSIONS","USERS","ENGAGEMENT %"],                       os_data,  [30,12,10,14]),
    ]
    for sec_title,headers,rows,widths in sections:
        ncols=len(headers)
        sheet_title(ws,row,ncols,f"{sec_title}  |  {DATE_START} -> {DATE_END}"); row+=1
        for ci,h in enumerate(headers,1): style_header(ws,row,ci,h)
        ws.row_dimensions[row].height=18; row+=1
        for rd in rows:
            bg=BG_DARK if row%2==1 else BG_MID
            for ci,val in enumerate(rd,1):
                style_data(ws,row,ci,val,bg=bg,fg=GOLD if ci>1 else WHITE,
                           align="center" if ci>1 else "left")
            ws.row_dimensions[row].height=16; row+=1
        row+=1
    for i,w in enumerate([30,12,10,14,18],1): set_col_width(ws,i,w)

# WRITE ALL SHEETS
def write_excel(data):
    ov=data["overview"]; pov=data["prev_overview"]
    os.makedirs(os.path.dirname(XLSX_PATH),exist_ok=True)
    wb=load_workbook(XLSX_PATH) if os.path.exists(XLSX_PATH) else Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]

    def fresh(name,pos=None):
        if name in wb.sheetnames: del wb[name]
        ws=wb.create_sheet(name,pos); ws.sheet_view.showGridLines=False; return ws

    # Dashboard
    ws_dash=fresh("Dashboard",0); build_dashboard(ws_dash,ov,pov)

    # History (append only)
    ws_hist=wb["History"] if "History" in wb.sheetnames else wb.create_sheet("History")
    build_history_sheet(ws_hist,ov)

    # Monthly
    ws_m=fresh("Monthly")
    build_table_sheet(ws_m,"MONTHLY TREND  |  LAST 12 MONTHS",
        ["MONTH","SESSIONS","USERS","NEW USERS","PAGE VIEWS","BOUNCE %","ENGAGEMENT %"],
        data["monthly"],[14,12,10,12,12,12,14],value_cols={2,3,4,5,6,7})
    if data["monthly"]:
        add_line(ws_dash,"Monthly Sessions",ws_m,3,3+len(data["monthly"])-1,1,2,"A13",w=36,h=12)

    # Weekly
    ws_wk=fresh("Weekly")
    build_table_sheet(ws_wk,"WEEKLY TREND  |  LAST 12 WEEKS",
        ["WEEK","SESSIONS","USERS","ENGAGEMENT %"],
        data["weekly"],[14,12,10,14],value_cols={2,3,4})
    if data["weekly"]:
        add_line(ws_dash,"Weekly Sessions",ws_wk,3,3+len(data["weekly"])-1,1,2,"A27",w=36,h=10)

    # Channels
    ws_ch=fresh("Channels")
    build_table_sheet(ws_ch,f"TRAFFIC CHANNELS  |  {DATE_START} -> {DATE_END}",
        ["CHANNEL","SESSIONS","USERS","NEW USERS","ENGAGEMENT %","BOUNCE %","AVG DUR (s)"],
        data["channels"],[28,12,10,12,14,12,18],value_cols={2,3,4,5,6,7})
    if data["channels"]:
        add_bar(ws_dash,"Sessions by Channel",ws_ch,3,3+len(data["channels"])-1,1,2,"G13",w=20,h=12)

    # Sources
    ws_src=fresh("Sources")
    build_table_sheet(ws_src,f"TRAFFIC SOURCES  |  {DATE_START} -> {DATE_END}",
        ["SOURCE","MEDIUM","SESSIONS","USERS","ENGAGEMENT %","AVG DUR (s)"],
        data["sources"],[30,20,12,10,14,18],value_cols={3,4,5,6})

    # Top Pages
    ws_pg=fresh("Top Pages")
    build_table_sheet(ws_pg,f"TOP PAGES  |  {DATE_START} -> {DATE_END}",
        ["PAGE PATH","VIEWS","USERS","AVG DUR (s)","BOUNCE %","ENGAGEMENT %"],
        data["pages"],[50,10,10,18,12,14],value_cols={2,3,4,5,6})
    if data["pages"]:
        add_bar(ws_dash,"Top Pages by Views",ws_pg,3,3+len(data["pages"])-1,1,2,"G27",w=20,h=10)

    # Landing Pages
    ws_lp=fresh("Landing Pages")
    build_table_sheet(ws_lp,f"LANDING PAGES  |  {DATE_START} -> {DATE_END}",
        ["LANDING PAGE","SESSIONS","USERS","BOUNCE %","ENGAGEMENT %","AVG DUR (s)"],
        data["landing"],[50,12,10,12,14,18],value_cols={2,3,4,5,6})

    # New vs Returning
    ws_nvr=fresh("New vs Returning")
    build_table_sheet(ws_nvr,f"NEW VS RETURNING  |  {DATE_START} -> {DATE_END}",
        ["TYPE","USERS","SESSIONS","ENGAGEMENT %","AVG DUR (s)","PAGES/SESSION"],
        data["new_vs_returning"],[16,10,12,14,18,14],value_cols={2,3,4,5,6})

    # Countries
    ws_cn=fresh("Countries")
    build_table_sheet(ws_cn,f"TOP COUNTRIES  |  {DATE_START} -> {DATE_END}",
        ["COUNTRY","USERS","SESSIONS","ENGAGEMENT %","AVG DUR (s)"],
        data["countries"],[30,10,12,14,18],value_cols={2,3,4,5})

    # Cities
    ws_ci=fresh("Cities")
    build_table_sheet(ws_ci,f"TOP CITIES  |  {DATE_START} -> {DATE_END}",
        ["CITY","COUNTRY","USERS","SESSIONS","ENGAGEMENT %"],
        data["cities"],[30,20,10,12,14],value_cols={3,4,5})

    # Technology
    ws_tech=fresh("Technology")
    build_technology_sheet(ws_tech,data["devices"],data["browsers"],data["os"])

    # Events
    ws_ev=fresh("Events")
    build_table_sheet(ws_ev,f"TOP EVENTS  |  {DATE_START} -> {DATE_END}",
        ["EVENT NAME","COUNT","USERS","COUNT / USER"],
        data["events"],[40,12,10,14],value_cols={2,3,4})

    # Day of Week
    ws_dow=fresh("Day of Week")
    build_table_sheet(ws_dow,f"TRAFFIC BY DAY OF WEEK  |  {DATE_START} -> {DATE_END}",
        ["DAY","SESSIONS","USERS","ENGAGEMENT %"],
        data["day_of_week"],[16,12,10,14],value_cols={2,3,4})
    if data["day_of_week"]:
        add_bar(ws_dow,"Sessions by Day",ws_dow,3,3+len(data["day_of_week"])-1,1,2,"F2",w=16,h=10)

    # Hour of Day
    ws_hod=fresh("Hour of Day")
    build_table_sheet(ws_hod,f"TRAFFIC BY HOUR  |  {DATE_START} -> {DATE_END}",
        ["HOUR","SESSIONS","ENGAGEMENT %"],
        data["hour_of_day"],[10,12,14],value_cols={2,3})
    if data["hour_of_day"]:
        add_line(ws_hod,"Sessions by Hour",ws_hod,3,3+len(data["hour_of_day"])-1,1,2,"E2",w=20,h=12)

    wb.save(XLSX_PATH)
    print(f"  Excel guardado em: {XLSX_PATH}")

# MAIN
if __name__=="__main__":
    try:
        print(f"\n{'='*62}")
        print(f"  EQC Google Analytics 4 Report  v2.0")
        print(f"  Periodo atual : {DATE_START}  ->  {DATE_END}  ({_days}d)")
        print(f"  Periodo prev  : {PREV_START}  ->  {PREV_END}")
        print(f"{'='*62}")

        if DATE_START>DATE_END:
            print("\n  Sem dados novos desde o ultimo run. Tenta amanha!")
            try:
                if sys.stdin and sys.stdin.isatty():
                    input("\nPrima ENTER para fechar...")
            except Exception:
                pass
            exit(0)

        client=get_client()
        print("\n  A carregar dados GA4...")

        print("  > Overview (atual + anterior)...")
        ov=fetch_overview(client,DATE_START,DATE_END)
        pov=fetch_overview(client,PREV_START,PREV_END)

        print("  > Paginas + Landing Pages...")
        pages=fetch_top_pages(client); landing=fetch_landing_pages(client)

        print("  > Canais + Fontes...")
        channels=fetch_channels(client); sources=fetch_sources(client)

        print("  > Paises + Cidades...")
        countries=fetch_countries(client); cities=fetch_cities(client)

        print("  > Dispositivos + Browsers + OS...")
        devices=fetch_devices(client); browsers=fetch_browsers(client); os_data=fetch_os(client)

        print("  > Eventos...")
        events=fetch_events(client)

        print("  > New vs Returning...")
        nvr=fetch_new_vs_returning(client)

        print("  > Tendencias (mensal, semanal, dia, hora)...")
        monthly=fetch_monthly(client); weekly=fetch_weekly(client)
        dow=fetch_day_of_week(client); hod=fetch_hour_of_day(client)

        def ds(c,p): return fmt_delta(c,p)[0]

        print(f"\n  {'─'*50}")
        print(f"  Sessoes       : {ov['sessions']:>6,}   {ds(ov['sessions'],pov['sessions'])}")
        print(f"  Utilizadores  : {ov['users']:>6,}   {ds(ov['users'],pov['users'])}")
        print(f"  Novos         : {ov['new_users']:>6,}   {ds(ov['new_users'],pov['new_users'])}")
        print(f"  Recorrentes   : {ov['returning_users']:>6,}   {ds(ov['returning_users'],pov['returning_users'])}")
        print(f"  Pag. vistas   : {ov['pageviews']:>6,}   {ds(ov['pageviews'],pov['pageviews'])}")
        print(f"  Dur. media    : {ov['avg_dur']:>6.0f}s  {ds(ov['avg_dur'],pov['avg_dur'])}")
        print(f"  Rejeicao      : {ov['bounce']:>5.1f}%   {ds(-ov['bounce'],-pov['bounce'])}")
        print(f"  Engagement    : {ov['engagement_rate']:>5.1f}%   {ds(ov['engagement_rate'],pov['engagement_rate'])}")
        print(f"  Sess./User    : {ov['sessions_per_user']:>6.2f}   {ds(ov['sessions_per_user'],pov['sessions_per_user'])}")
        print(f"  Pag./Sessao   : {ov['pages_per_session']:>6.2f}   {ds(ov['pages_per_session'],pov['pages_per_session'])}")
        print(f"  {'─'*50}")

        print("\n  A gerar Excel (16 sheets)...")
        write_excel({
            "overview":ov,"prev_overview":pov,
            "pages":pages,"landing":landing,
            "channels":channels,"sources":sources,
            "countries":countries,"cities":cities,
            "devices":devices,"browsers":browsers,"os":os_data,
            "events":events,"new_vs_returning":nvr,
            "monthly":monthly,"weekly":weekly,
            "day_of_week":dow,"hour_of_day":hod,
        })

        print(f"\n{'='*62}")
        print("  Concluido!  16 sheets geradas.")
        print(f"{'='*62}\n")

    except Exception as e:
        import traceback
        print(f"\n[ERRO] {type(e).__name__}: {e}")
        traceback.print_exc()
        print("\nVerifica:")
        print("  1. credentials.json existe em C:\\eqc\\")
        print("  2. pip install google-analytics-data openpyxl")
        print("  3. Service Account tem acesso ao GA4 como Viewer")

    # Só pede ENTER se houver consola interactiva (double-click / cmd manual).
    # Corrida pelo Task Scheduler via pythonw não tem stdin — input() bloquearia
    # ou rebentaria para sempre, e a tarefa nunca terminaria.
    try:
        if sys.stdin and sys.stdin.isatty():
            input("\nPrima ENTER para fechar...")
    except Exception:
        pass
